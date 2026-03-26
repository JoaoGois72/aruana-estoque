from datetime import datetime
from decimal import Decimal
from io import BytesIO

from flask import render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from sqlalchemy import and_, or_

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from openpyxl import Workbook

from app.extensions import db
from app.models.material import Material
from app.models.solicitacao import Solicitacao
from app.models.solicitacao_item import SolicitacaoItem
from app.models.entrada import Entrada
from app.models.categoria import Categoria

# Se você tiver Fornecedor no projeto, descomente:
# from app.models.fornecedor import Fornecedor

from app.blueprints.relatorios import relatorios_bp
from app.blueprints.estoque import estoque_bp

def role_required(*roles):
    from functools import wraps
    from flask_login import current_user
    from flask import abort, redirect

    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):

            if not current_user.is_authenticated:
                return redirect("/auth/login")

            if roles and current_user.role not in roles:
                abort(403)

            return fn(*args, **kwargs)

        return wrapper
    return decorator
# =========================
# Helpers
# =========================
def _parse_date(s: str | None):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None


def _d(v) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    return Decimal(str(v))


def _wb_to_bytes(wb: Workbook) -> BytesIO:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _pdf_table(title: str, headers: list[str], rows: list[list[str]], filename: str):
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    x = 15 * mm
    y = h - 20 * mm

    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, title)
    y -= 10 * mm

    c.setFont("Helvetica", 9)
    c.drawString(x, y, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    y -= 8 * mm

    # Cabeçalho
    c.setFont("Helvetica-Bold", 9)
    colw = (w - 30 * mm) / max(1, len(headers))
    for i, head in enumerate(headers):
        c.drawString(x + i * colw, y, head[:28])
    y -= 6 * mm

    c.setFont("Helvetica", 9)
    for row in rows:
        if y < 20 * mm:
            c.showPage()
            y = h - 20 * mm
            c.setFont("Helvetica-Bold", 9)
            for i, head in enumerate(headers):
                c.drawString(x + i * colw, y, head[:28])
            y -= 6 * mm
            c.setFont("Helvetica", 9)

        for i, cell in enumerate(row):
            c.drawString(x + i * colw, y, str(cell)[:28])
        y -= 5 * mm

    c.showPage()
    c.save()
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


# =========================
# Index
# =========================
@relatorios_bp.get("/")
@login_required
def relatorios_index():
    return render_template("relatorios/index.html")


# =========================
# 1) ESTOQUE ATUAL
# =========================
@relatorios_bp.get("/estoque")
@login_required
def relatorio_estoque():
    q = request.args.get("q", "").strip()

    query = Material.query.filter_by(ativo=True)
    if q:
        query = query.filter(Material.nome.ilike(f"%{q}%"))

    materiais = query.order_by(Material.nome.asc()).all()
    return render_template("relatorios/estoque.html", materiais=materiais, q=q)


@relatorios_bp.get("/estoque.xlsx")
@login_required
def relatorio_estoque_xlsx():
    q = request.args.get("q", "").strip()

    query = Material.query.filter_by(ativo=True)
    if q:
        query = query.filter(Material.nome.ilike(f"%{q}%"))
    materiais = query.order_by(Material.nome.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"
    ws.append(["Código", "Nome", "Unidade", "Saldo", "Reservado", "Disponível", "Mínimo"])

    for m in materiais:
        saldo = _d(getattr(m, "saldo_atual", 0))
        reservado = _d(getattr(m, "reservado_atual", 0))
        minimo = _d(getattr(m, "estoque_minimo", 0))
        disponivel = saldo - reservado
        ws.append([
            m.codigo or "",
            m.nome,
            m.unidade,
            float(saldo),
            float(reservado),
            float(disponivel),
            float(minimo),
        ])

    bio = _wb_to_bytes(wb)
    return send_file(
        bio,
        as_attachment=True,
        download_name="relatorio_estoque.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@relatorios_bp.get("/estoque.pdf")
@login_required
def relatorio_estoque_pdf():
    q = request.args.get("q", "").strip()
    query = Material.query.filter_by(ativo=True)
    if q:
        query = query.filter(Material.nome.ilike(f"%{q}%"))
    materiais = query.order_by(Material.nome.asc()).all()

    headers = ["Código", "Nome", "Un", "Saldo", "Res", "Disp", "Min"]
    rows = []
    for m in materiais:
        saldo = _d(getattr(m, "saldo_atual", 0))
        reservado = _d(getattr(m, "reservado_atual", 0))
        minimo = _d(getattr(m, "estoque_minimo", 0))
        disponivel = saldo - reservado
        rows.append([
            m.codigo or "-",
            m.nome,
            m.unidade,
            str(saldo),
            str(reservado),
            str(disponivel),
            str(minimo),
        ])

    return _pdf_table("Relatório de Estoque Atual", headers, rows, "relatorio_estoque.pdf")


# =========================
# 2) CONSUMO (ENTREGUE) POR TORRE/APTO
# =========================
@relatorios_bp.get("/consumo")
@login_required
def relatorio_consumo():
    data_de = _parse_date(request.args.get("de"))
    data_ate = _parse_date(request.args.get("ate"))
    torre = (request.args.get("torre") or "").strip()
    pav = (request.args.get("pav") or "").strip()
    apto = (request.args.get("apto") or "").strip()

    filtros = [Solicitacao.status == "ENTREGUE"]
    if data_de:
        filtros.append(Solicitacao.data_entrega >= data_de)
    if data_ate:
        filtros.append(Solicitacao.data_entrega <= data_ate)
    if torre:
        filtros.append(Solicitacao.local_torre == torre)
    if pav:
        filtros.append(Solicitacao.local_pav == pav)
    if apto:
        filtros.append(Solicitacao.local_apto == apto)

    itens = (
        SolicitacaoItem.query
        .join(Solicitacao, Solicitacao.id == SolicitacaoItem.solicitacao_id)
        .options(joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .all()
    )

    return render_template(
        "relatorios/consumo.html",
        itens=itens,
        de=request.args.get("de", ""),
        ate=request.args.get("ate", ""),
        torre=torre,
        pav=pav,
        apto=apto,
    )


@relatorios_bp.get("/consumo.xlsx")
@login_required
def relatorio_consumo_xlsx():
    data_de = _parse_date(request.args.get("de"))
    data_ate = _parse_date(request.args.get("ate"))
    torre = (request.args.get("torre") or "").strip()
    pav = (request.args.get("pav") or "").strip()
    apto = (request.args.get("apto") or "").strip()

    filtros = [Solicitacao.status == "ENTREGUE"]
    if data_de:
        filtros.append(Solicitacao.data_entrega >= data_de)
    if data_ate:
        filtros.append(Solicitacao.data_entrega <= data_ate)
    if torre:
        filtros.append(Solicitacao.local_torre == torre)
    if pav:
        filtros.append(Solicitacao.local_pav == pav)
    if apto:
        filtros.append(Solicitacao.local_apto == apto)

    itens = (
        SolicitacaoItem.query
        .join(Solicitacao, Solicitacao.id == SolicitacaoItem.solicitacao_id)
        .options(joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo"
    ws.append(["Solic#", "Entrega", "Torre", "Pav", "Apto", "Material", "Qtd", "Un"])

    for it in itens:
        s = it.solicitacao  # relationship (se você não tiver, me diga)
        m = it.material
        ws.append([
            s.id,
            s.data_entrega.strftime("%d/%m/%Y") if s.data_entrega else "",
            s.local_torre or "",
            s.local_pav or "",
            s.local_apto or "",
            m.nome if m else "",
            float(_d(it.qtd)),
            m.unidade if m else "",
        ])

    bio = _wb_to_bytes(wb)
    return send_file(
        bio,
        as_attachment=True,
        download_name="relatorio_consumo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@relatorios_bp.get("/consumo.pdf")
@login_required
def relatorio_consumo_pdf():
    data_de = _parse_date(request.args.get("de"))
    data_ate = _parse_date(request.args.get("ate"))
    torre = (request.args.get("torre") or "").strip()
    pav = (request.args.get("pav") or "").strip()
    apto = (request.args.get("apto") or "").strip()

    filtros = [Solicitacao.status == "ENTREGUE"]
    if data_de:
        filtros.append(Solicitacao.data_entrega >= data_de)
    if data_ate:
        filtros.append(Solicitacao.data_entrega <= data_ate)
    if torre:
        filtros.append(Solicitacao.local_torre == torre)
    if pav:
        filtros.append(Solicitacao.local_pav == pav)
    if apto:
        filtros.append(Solicitacao.local_apto == apto)

    itens = (
        SolicitacaoItem.query
        .join(Solicitacao, Solicitacao.id == SolicitacaoItem.solicitacao_id)
        .options(joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .all()
    )

    headers = ["Solic#", "Entrega", "Local", "Material", "Qtd", "Un"]
    rows = []
    for it in itens:
        s = it.solicitacao
        m = it.material
        local = f"{s.local_torre or ''}-{s.local_pav or ''}-{s.local_apto or ''}"
        rows.append([
            str(s.id),
            s.data_entrega.strftime("%d/%m/%Y") if s.data_entrega else "",
            local,
            m.nome if m else "",
            str(_d(it.qtd)),
            m.unidade if m else "",
        ])

    return _pdf_table("Relatório de Consumo (ENTREGUE)", headers, rows, "relatorio_consumo.pdf")


# =========================
# 3) ENTRADAS POR FORNECEDOR
# =========================
@relatorios_bp.get("/entradas")
@login_required
def relatorio_entradas_fornecedor():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))
    doc = (request.args.get("doc") or "").strip()

    q = Entrada.query
    if de:
        q = q.filter(Entrada.data_entrada >= de)
    if ate:
        q = q.filter(Entrada.data_entrada <= ate)
    if doc:
        # Se você armazena documento do fornecedor na entrada:
        # q = q.filter(Entrada.documento_fornecedor == doc)
        # Se não tiver, filtramos pelo texto (nome_fornecedor):
        q = q.filter(Entrada.nome_fornecedor.ilike(f"%{doc}%"))

    entradas = q.order_by(Entrada.id.desc()).limit(500).all()

    return render_template(
        "relatorios/entradas.html",
        entradas=entradas,
        de=request.args.get("de", ""),
        ate=request.args.get("ate", ""),
        doc=doc,
    )


@relatorios_bp.get("/entradas.xlsx")
@login_required
def relatorio_entradas_fornecedor_xlsx():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))
    doc = (request.args.get("doc") or "").strip()

    q = Entrada.query
    if de:
        q = q.filter(Entrada.data_entrada >= de)
    if ate:
        q = q.filter(Entrada.data_entrada <= ate)
    if doc:
        q = q.filter(Entrada.nome_fornecedor.ilike(f"%{doc}%"))

    entradas = q.order_by(Entrada.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Entradas"
    ws.append(["Entrada#", "Data", "NF", "Fornecedor", "Status"])

    for e in entradas:
        ws.append([
            e.id,
            e.data_entrada.strftime("%d/%m/%Y") if e.data_entrada else "",
            getattr(e, "numero_nf", "") or "",
            getattr(e, "nome_fornecedor", "") or "",
            getattr(e, "status", "") or "",
        ])

    bio = _wb_to_bytes(wb)
    return send_file(
        bio,
        as_attachment=True,
        download_name="relatorio_entradas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@relatorios_bp.get("/entradas.pdf")
@login_required
def relatorio_entradas_fornecedor_pdf():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))
    doc = (request.args.get("doc") or "").strip()

    q = Entrada.query
    if de:
        q = q.filter(Entrada.data_entrada >= de)
    if ate:
        q = q.filter(Entrada.data_entrada <= ate)
    if doc:
        q = q.filter(Entrada.nome_fornecedor.ilike(f"%{doc}%"))

    entradas = q.order_by(Entrada.id.desc()).all()

    headers = ["Entrada#", "Data", "NF", "Fornecedor", "Status"]
    rows = []
    for e in entradas:
        rows.append([
            str(e.id),
            e.data_entrada.strftime("%d/%m/%Y") if e.data_entrada else "",
            getattr(e, "numero_nf", "") or "",
            getattr(e, "nome_fornecedor", "") or "",
            getattr(e, "status", "") or "",
        ])

    return _pdf_table("Relatório de Entradas por Fornecedor", headers, rows, "relatorio_entradas.pdf")


# =========================
# 4) SAÍDAS POR PERÍODO (ENTREGUE)
# =========================
@relatorios_bp.get("/saidas")
@login_required
def relatorio_saidas_periodo():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))

    filtros = [Solicitacao.status == "ENTREGUE"]
    if de:
        filtros.append(Solicitacao.data_entrega >= de)
    if ate:
        filtros.append(Solicitacao.data_entrega <= ate)

    solicitacoes = (
        Solicitacao.query
        .options(joinedload(Solicitacao.itens).joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .limit(300)
        .all()
    )

    return render_template(
        "relatorios/saidas.html",
        solicitacoes=solicitacoes,
        de=request.args.get("de", ""),
        ate=request.args.get("ate", ""),
    )


@relatorios_bp.get("/saidas.xlsx")
@login_required
def relatorio_saidas_periodo_xlsx():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))

    filtros = [Solicitacao.status == "ENTREGUE"]
    if de:
        filtros.append(Solicitacao.data_entrega >= de)
    if ate:
        filtros.append(Solicitacao.data_entrega <= ate)

    solicitacoes = (
        Solicitacao.query
        .options(joinedload(Solicitacao.itens).joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Saídas"
    ws.append(["Solic#", "Entrega", "Torre", "Pav", "Apto", "Material", "Qtd", "Un"])

    for s in solicitacoes:
        for it in s.itens:
            m = it.material
            ws.append([
                s.id,
                s.data_entrega.strftime("%d/%m/%Y") if s.data_entrega else "",
                s.local_torre or "",
                s.local_pav or "",
                s.local_apto or "",
                m.nome if m else "",
                float(_d(it.qtd)),
                m.unidade if m else "",
            ])

    bio = _wb_to_bytes(wb)
    return send_file(
        bio,
        as_attachment=True,
        download_name="relatorio_saidas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@relatorios_bp.get("/saidas.pdf")
@login_required
def relatorio_saidas_periodo_pdf():
    de = _parse_date(request.args.get("de"))
    ate = _parse_date(request.args.get("ate"))

    filtros = [Solicitacao.status == "ENTREGUE"]
    if de:
        filtros.append(Solicitacao.data_entrega >= de)
    if ate:
        filtros.append(Solicitacao.data_entrega <= ate)

    solicitacoes = (
        Solicitacao.query
        .options(joinedload(Solicitacao.itens).joinedload(SolicitacaoItem.material))
        .filter(and_(*filtros))
        .order_by(Solicitacao.id.desc())
        .all()
    )

    headers = ["Solic#", "Entrega", "Local", "Material", "Qtd", "Un"]
    rows = []
    for s in solicitacoes:
        local = f"{s.local_torre or ''}-{s.local_pav or ''}-{s.local_apto or ''}"
        for it in s.itens:
            m = it.material
            rows.append([
                str(s.id),
                s.data_entrega.strftime("%d/%m/%Y") if s.data_entrega else "",
                local,
                m.nome if m else "",
                str(_d(it.qtd)),
                m.unidade if m else "",
            ])

    return _pdf_table("Relatório de Saídas (ENTREGUE) por Período", headers, rows, "relatorio_saidas.pdf")

# =========================
# LISTAR MATERIAIS
# =========================
@estoque_bp.route("/materiais")
@login_required
def materiais():

    busca = request.args.get("busca", "")
    categoria_id = int(request.form.get("categoria_id") or 0)

    query = Material.query.filter_by(ativo=True)

    if busca:
        query = query.filter(
            or_(
                Material.nome.ilike(f"%{busca}%"),
                Material.codigo.ilike(f"%{busca}%")
            )
        )

    if categoria_id:
        query = query.filter(Material.categoria_id == categoria_id)

    materiais = query.order_by(Material.nome).all()
    categorias = Categoria.query.order_by(Categoria.nome).all()

    return render_template(
        "estoque/materiais.html",
        materiais=materiais,
        categorias=categorias,
        busca=busca,
        categoria_id=categoria_id
    )


# =========================
# NOVO MATERIAL
# =========================
from decimal import Decimal

@estoque_bp.route("/materiais/novo", methods=["GET", "POST"])
@login_required
def material_novo():

    if request.method == "POST":

        codigo = request.form.get("codigo")
        nome = request.form.get("nome")
        unidade = request.form.get("unidade")
        categoria_id = request.form.get("categoria_id")

        estoque_minimo = Decimal(request.form.get("estoque_minimo") or "0")
        saldo_atual = Decimal(request.form.get("saldo_atual") or "0")

        # 🚨 VALIDAÇÃO BÁSICA
        if not nome:
            flash("Informe a descrição do material", "danger")
            return redirect(url_for("estoque.material_novo"))

        # 🔥 GERAR CÓDIGO AUTOMÁTICO (100% seguro)
        if not codigo or codigo.strip() == "":

            ultimo = Material.query.order_by(Material.id.desc()).first()

            if ultimo and ultimo.codigo and ultimo.codigo.isdigit():
                codigo = str(int(ultimo.codigo) + 1)
            else:
                codigo = "1"

        # 🔥 GARANTIR FORMATO
        codigo = f"{int(codigo):04d}"

        # 🚨 EVITAR DUPLICADO
        existe = Material.query.filter_by(codigo=codigo).first()
        if existe:
            flash(f"Código {codigo} já existe!", "danger")
            return redirect(url_for("estoque.material_novo"))

        material = Material(
            codigo=codigo,
            nome=nome,
            unidade=unidade,
            estoque_minimo=estoque_minimo,
            saldo_atual=saldo_atual,
            categoria_id=int(categoria_id) if categoria_id else None,
            ativo=True
        )

        db.session.add(material)
        db.session.commit()

        flash(f"Material {nome} cadastrado!", "success")
        return redirect(url_for("estoque.materiais"))

    categorias = Categoria.query.order_by(Categoria.nome).all()

    return render_template(
        "estoque/material_form.html",
        categorias=categorias,
        material=None
    )

# =========================
# EDITAR MATERIAL
# =========================
@estoque_bp.route("/materiais/<int:id>/editar", methods=["GET", "POST"])
@login_required
def material_editar(id):

    material = Material.query.get_or_404(id)
    categorias = Categoria.query.order_by(Categoria.nome).all()

    if request.method == "POST":

        material.codigo = request.form.get("codigo")
        material.nome = request.form.get("nome")
        material.unidade = request.form.get("unidade")
        material.categoria_id = request.form.get("categoria_id")
        material.estoque_minimo = request.form.get("estoque_minimo") or 0
        material.saldo_atual = request.form.get("saldo_atual") or 0

        db.session.commit()

        flash("Material atualizado com sucesso!", "success")
        return redirect(url_for("estoque.materiais"))

    return render_template(
        "estoque/material_form.html",
        material=material,
        categorias=categorias
    )


# =========================
# INATIVAR MATERIAL
# =========================
@estoque_bp.route("/materiais/<int:id>/inativar")
@login_required
def material_inativar(id):

    material = Material.query.get_or_404(id)

    material.ativo = False
    db.session.commit()

    flash("Material inativado!", "warning")
    return redirect(url_for("estoque.materiais"))

