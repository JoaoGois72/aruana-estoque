from datetime import datetime, time
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from sqlalchemy.orm import joinedload

from app.models.material import Material
from app.models.solicitacao import Solicitacao
from app.models.solicitacao_item import SolicitacaoItem
from app.models.user import User


STATUS_SOLICITACOES = [
    "PENDENTE",
    "ANALISE_PARCIAL",
    "APROVADA",
    "APROVADA_PARCIAL",
    "REJEITADA",
    "ENTREGUE",
    "ENTREGUE_PARCIAL",
]


def converter_data(valor, final_do_dia=False):
    if not valor:
        return None

    try:
        data = datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("Data informada em formato inválido.")

    if final_do_dia:
        return datetime.combine(data, time.max)

    return datetime.combine(data, time.min)


def montar_query_solicitacoes(filtros, current_user):
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.aprovado_por),
            joinedload(Solicitacao.entregue_por),
            joinedload(Solicitacao.itens)
            .joinedload(SolicitacaoItem.material),
        )
    )

    perfis_acesso_total = {
        "ADMIN",
        "ENGENHEIRO",
        "ALMOXARIFE",
        "AUX_ALMOX",
    }

    if current_user.role not in perfis_acesso_total:
        query = query.filter(
            Solicitacao.usuario_id == current_user.id
        )

    status = (filtros.get("status") or "").strip()

    if status:
        query = query.filter(
            Solicitacao.status == status
        )

    usuario_id = filtros.get("usuario_id")

    if usuario_id:
        try:
            usuario_id = int(usuario_id)
        except (TypeError, ValueError):
            raise ValueError("Solicitante inválido.")

        query = query.filter(
            Solicitacao.usuario_id == usuario_id
        )

    torre = (filtros.get("torre") or "").strip()

    if torre:
        query = query.filter(
            Solicitacao.local_torre == torre
        )

    pavimento = (filtros.get("pavimento") or "").strip()

    if pavimento:
        query = query.filter(
            Solicitacao.local_pav == pavimento
        )

    apartamento = (
        filtros.get("apartamento") or ""
    ).strip()

    if apartamento:
        query = query.filter(
            Solicitacao.local_apto == apartamento
        )

    material_id = filtros.get("material_id")

    if material_id:
        try:
            material_id = int(material_id)
        except (TypeError, ValueError):
            raise ValueError("Material inválido.")

        query = (
            query
            .join(
                SolicitacaoItem,
                SolicitacaoItem.solicitacao_id
                == Solicitacao.id,
            )
            .filter(
                SolicitacaoItem.material_id == material_id
            )
            .distinct()
        )

    data_inicial = converter_data(
        filtros.get("data_inicial")
    )

    if data_inicial:
        query = query.filter(
            Solicitacao.data_solicitacao >= data_inicial
        )

    data_final = converter_data(
        filtros.get("data_final"),
        final_do_dia=True,
    )

    if data_final:
        query = query.filter(
            Solicitacao.data_solicitacao <= data_final
        )

    return query.order_by(
        Solicitacao.data_solicitacao.desc(),
        Solicitacao.id.desc(),
    )


def listar_solicitacoes(filtros, current_user):
    return montar_query_solicitacoes(
        filtros,
        current_user,
    ).all()


def decimal_para_float(valor):
    if valor is None:
        return None

    return float(Decimal(valor))


def formatar_decimal(valor):
    if valor is None:
        return "-"

    numero = Decimal(valor)

    return (
        f"{numero:,.2f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def formatar_data(valor):
    if not valor:
        return "-"

    return valor.strftime("%d/%m/%Y %H:%M")


def formatar_local(solicitacao):
    partes = [
        solicitacao.local_torre,
        solicitacao.local_pav,
        solicitacao.local_apto,
    ]

    partes = [
        str(parte).strip()
        for parte in partes
        if parte and str(parte).strip()
    ]

    return " / ".join(partes) if partes else "-"


def montar_linhas_relatorio(solicitacoes):
    linhas = []

    for solicitacao in solicitacoes:
        if not solicitacao.itens:
            linhas.append({
                "solicitacao_id": solicitacao.id,
                "data_solicitacao": solicitacao.data_solicitacao,
                "solicitante": (
                    solicitacao.usuario.nome
                    if solicitacao.usuario
                    else "-"
                ),
                "local": formatar_local(solicitacao),
                "status_solicitacao": solicitacao.status,
                "material": "-",
                "codigo_material": "-",
                "unidade": "-",
                "qtd_solicitada": None,
                "qtd_aprovada": None,
                "status_item": "-",
                "motivo_rejeicao": "-",
                "aprovado_por": (
                    solicitacao.aprovado_por.nome
                    if solicitacao.aprovado_por
                    else "-"
                ),
                "entregue_por": (
                    solicitacao.entregue_por.nome
                    if solicitacao.entregue_por
                    else "-"
                ),
                "data_aprovacao": solicitacao.data_aprovacao,
                "data_entrega": solicitacao.data_entrega,
            })

            continue

        for item in solicitacao.itens:
            linhas.append({
                "solicitacao_id": solicitacao.id,
                "data_solicitacao": solicitacao.data_solicitacao,
                "solicitante": (
                    solicitacao.usuario.nome
                    if solicitacao.usuario
                    else "-"
                ),
                "local": formatar_local(solicitacao),
                "status_solicitacao": solicitacao.status,
                "material": (
                    item.material.nome
                    if item.material
                    else "-"
                ),
                "codigo_material": (
                    item.material.codigo
                    if item.material
                    else "-"
                ),
                "unidade": (
                    item.material.unidade
                    if item.material
                    else "-"
                ),
                "qtd_solicitada": item.qtd,
                "qtd_aprovada": item.qtd_aprovada,
                "status_item": item.status,
                "motivo_rejeicao": (
                    item.motivo_rejeicao or "-"
                ),
                "aprovado_por": (
                    solicitacao.aprovado_por.nome
                    if solicitacao.aprovado_por
                    else "-"
                ),
                "entregue_por": (
                    solicitacao.entregue_por.nome
                    if solicitacao.entregue_por
                    else "-"
                ),
                "data_aprovacao": solicitacao.data_aprovacao,
                "data_entrega": solicitacao.data_entrega,
            })

    return linhas
  def gerar_excel_solicitacoes(solicitacoes):
    linhas = montar_linhas_relatorio(solicitacoes)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Solicitações"

    cabecalhos = [
        "Solicitação",
        "Data da solicitação",
        "Solicitante",
        "Local",
        "Status da solicitação",
        "Código",
        "Material",
        "Unidade",
        "Qtd. solicitada",
        "Qtd. aprovada",
        "Status do item",
        "Motivo da rejeição",
        "Aprovado por",
        "Data da aprovação",
        "Entregue por",
        "Data da entrega",
    ]

    planilha.append(cabecalhos)

    preenchimento = PatternFill(
        fill_type="solid",
        fgColor="0D6EFD",
    )

    for celula in planilha[1]:
        celula.font = Font(
            bold=True,
            color="FFFFFF",
        )
        celula.fill = preenchimento
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for linha in linhas:
        planilha.append([
            linha["solicitacao_id"],
            formatar_data(linha["data_solicitacao"]),
            linha["solicitante"],
            linha["local"],
            linha["status_solicitacao"],
            linha["codigo_material"],
            linha["material"],
            linha["unidade"],
            decimal_para_float(
                linha["qtd_solicitada"]
            ),
            decimal_para_float(
                linha["qtd_aprovada"]
            ),
            linha["status_item"],
            linha["motivo_rejeicao"],
            linha["aprovado_por"],
            formatar_data(linha["data_aprovacao"]),
            linha["entregue_por"],
            formatar_data(linha["data_entrega"]),
        ])

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions

    larguras = {
        "A": 12,
        "B": 20,
        "C": 25,
        "D": 25,
        "E": 22,
        "F": 14,
        "G": 40,
        "H": 12,
        "I": 16,
        "J": 16,
        "K": 18,
        "L": 35,
        "M": 25,
        "N": 20,
        "O": 25,
        "P": 20,
    }

    for coluna, largura in larguras.items():
        planilha.column_dimensions[coluna].width = largura

    for linha in planilha.iter_rows(
        min_row=2,
        max_row=planilha.max_row,
    ):
        for celula in linha:
            celula.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for coluna in ["I", "J"]:
        for celula in planilha[coluna][1:]:
            celula.number_format = '#,##0.00'

    resumo = workbook.create_sheet("Resumo")

    resumo.append([
        "Indicador",
        "Quantidade",
    ])

    for celula in resumo[1]:
        celula.font = Font(
            bold=True,
            color="FFFFFF",
        )
        celula.fill = preenchimento

    totais_status = {}

    for solicitacao in solicitacoes:
        totais_status[solicitacao.status] = (
            totais_status.get(
                solicitacao.status,
                0,
            )
            + 1
        )

    resumo.append([
        "Total de solicitações",
        len(solicitacoes),
    ])

    for status in STATUS_SOLICITACOES:
        resumo.append([
            status.replace("_", " ").title(),
            totais_status.get(status, 0),
        ])

    resumo.column_dimensions["A"].width = 35
    resumo.column_dimensions["B"].width = 18

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)

    return arquivo

def gerar_pdf_solicitacoes(
    solicitacoes,
    filtros,
):
    linhas = montar_linhas_relatorio(solicitacoes)

    arquivo = BytesIO()

    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=0.6 * cm,
        leftMargin=0.6 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
        title="Relatório de Solicitações",
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloRelatorio",
        parent=estilos["Title"],
        fontSize=16,
        leading=19,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    estilo_normal = ParagraphStyle(
        "TextoTabela",
        parent=estilos["BodyText"],
        fontSize=6.5,
        leading=8,
    )

    elementos = [
        Paragraph(
            "Aruana Garden – Relatório de Solicitações",
            estilo_titulo,
        )
    ]

    periodo = "Todos os períodos"

    if filtros.get("data_inicial") or filtros.get("data_final"):
        periodo = (
            f"{filtros.get('data_inicial') or 'Início'}"
            f" até "
            f"{filtros.get('data_final') or 'Hoje'}"
        )

    descricao_filtros = (
        f"Período: {periodo} | "
        f"Status: {filtros.get('status') or 'Todos'} | "
        f"Total de solicitações: {len(solicitacoes)}"
    )

    elementos.append(
        Paragraph(
            descricao_filtros,
            estilos["Normal"],
        )
    )

    elementos.append(
        Spacer(1, 0.3 * cm)
    )

    dados = [[
        "Nº",
        "Data",
        "Solicitante",
        "Local",
        "Status",
        "Material",
        "Qtd. sol.",
        "Qtd. apr.",
        "Status item",
        "Motivo",
    ]]

    for linha in linhas:
        dados.append([
            str(linha["solicitacao_id"]),
            Paragraph(
                formatar_data(
                    linha["data_solicitacao"]
                ),
                estilo_normal,
            ),
            Paragraph(
                linha["solicitante"],
                estilo_normal,
            ),
            Paragraph(
                linha["local"],
                estilo_normal,
            ),
            Paragraph(
                linha["status_solicitacao"],
                estilo_normal,
            ),
            Paragraph(
                (
                    f"{linha['codigo_material']} - "
                    f"{linha['material']}"
                ),
                estilo_normal,
            ),
            formatar_decimal(
                linha["qtd_solicitada"]
            ),
            formatar_decimal(
                linha["qtd_aprovada"]
            ),
            Paragraph(
                linha["status_item"],
                estilo_normal,
            ),
            Paragraph(
                linha["motivo_rejeicao"],
                estilo_normal,
            ),
        ])

    tabela = Table(
        dados,
        repeatRows=1,
        colWidths=[
            1.0 * cm,
            2.2 * cm,
            3.0 * cm,
            3.0 * cm,
            2.5 * cm,
            5.2 * cm,
            1.6 * cm,
            1.6 * cm,
            2.3 * cm,
            4.2 * cm,
        ],
    )

    tabela.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#0D6EFD"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                6.5,
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, 0),
                "CENTER",
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.25,
                colors.HexColor("#CED4DA"),
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#F8F9FA"),
                ],
            ),
            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                3,
            ),
        ])
    )

    elementos.append(tabela)

    documento.build(elementos)

    arquivo.seek(0)

    return arquivo
  
